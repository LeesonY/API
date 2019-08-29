# -*- coding: utf-8 -*-
# @Time : 2019/3/10 16:35
# @Author : 小雨点
# @Email: 1149750734@qq.com
# @File : do_excel.py

from openpyxl import load_workbook
from pythonLX.week10.API_5.common import project_path
from pythonLX.week10.API_5.common.read_config import ReadConfig

class DoExcel:
    '''该类完成数据读取和写入'''
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name#Excel工作簿文件名或地址
        self.sheet_name=sheet_name#表单名

    def readexcel(self,section):#section配置文件中的片断名，可以根据指定来执行用例
        '''从Excel读取数据'''
        #从配置文件控制读取哪条用例
        case_id=ReadConfig(project_path.conf_path).get_data(section,'case_id')
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]

        '''读取数据,获取存在excel中的tel'''
        tel=self.get_tel()#类里面的方法调用，self点+方法名
        test_data=[]
        for i in range(2,sheet.max_row+1):
            row_data={}
            row_data['CaseId']=sheet.cell(i,1).value
            row_data['Module']=sheet.cell(i,2).value
            row_data['Title']=sheet.cell(i,3).value
            row_data['Url']=sheet.cell(i,4).value
            row_data['Method']=sheet.cell(i,5).value
            '''(手机号码参数化)find找到返回，找不到返回-1'''
            if sheet.cell(i,6).value.find('tel')!=-1:
                '''#把Params的tel替换成13601676495'''
                row_data['Params']=sheet.cell(i,6).value.replace('tel',str(tel))#替换tel,str(tel)因为replace是字符串的替换。
                self.update_tel(int(tel)+1)#写入excel
            else:#如果没有tel，就不用替换
                row_data['Params']=sheet.cell(i,6).value
            row_data['Sql']=sheet.cell(i,7).value#读取sql语句
            row_data['ExpectedResult']=sheet.cell(i,8).value
            test_data.append(row_data)
        wb.close()
        final_data=[]#空列表 存储最终的测试用例数据
        if case_id=='all':#如果case_id==all 那就获取所有的用例数据
           final_data=test_data#把测试用例赋值给final_data这个变量
        else:#如果是列表 那就获取列表里面指定id的用例的数据
           for i in case_id:#遍历case_id 里面的值
               final_data.append(test_data[i-1])
        return final_data
    def get_tel(self):
        '''获取excel里面保存的手机号'''
        wb=load_workbook(self.file_name)#打开excel
        sheet=wb['tel']#表单名写死
        wb.close()
        return sheet.cell(1,2).value#返回tel的值
    def update_tel(self,new_tel):
        '''写回电话号码'''
        wb=load_workbook(self.file_name)#打开excel
        sheet=wb['tel']#表单名写死
        sheet.cell(1,2,new_tel)#写入新的电话号码
        wb.save(self.file_name)
        wb.close()
    def write_back(self,row,col,value):
        '''写会excel'''
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]

        sheet.cell(row,col).value=value

        wb.save(self.file_name)
        wb.close()#关闭文件

if __name__ == '__main__':
    # test_data=DoExcel(project_path.case_path,'register').readexcel()
    test_data=DoExcel(project_path.case_path,'recharge').readexcel('RechargeCASE')
    print(test_data)

    # tel=DoExcel(project_path.case_path,'tel').get_tel()
    # print(tel)
