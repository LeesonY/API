# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from API_3.common.read_config import ReadConfig
from API_3.common import project_path



class DoExcel:
    '''该类完成测试数据的读取以及测试结果的写回'''

    def __init__(self, file_name, sheet_name):
        self.file_name = file_name  # Excel工作簿文件名或地址
        self.sheet_name = sheet_name  # 表单名

    def read_data(self):
        '''从Excel读取数据，有返回值'''
        #从配置文件里面控制读取哪些用例
        case_id =ReadConfig(project_path.conf_path).get_data('CASE','Case_id')
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]

        # 唯一的要求是什么？每一行数据要在一起{} [] ()
        # 如何把每一行的数据存在一个空间里面去?[]
        # 开始读取数据
        # 获取存在Excel里面的手机号码
        tel=self.get_tel()#自己传递表单进来
        test_data = []
        for i in range(2, sheet.max_row + 1):
            row_data = {}
            row_data['CaseId'] = sheet.cell(i, 1).value
            row_data['Module'] = sheet.cell(i, 2).value
            row_data['Title'] = sheet.cell(i, 3).value
            row_data['Url'] = sheet.cell(i, 4).value
            row_data['Method'] = sheet.cell(i, 5).value
            if sheet.cell(i, 6).value.find('tel')!=-1:#注意这个方法的使用以及返回值 也可以用成员运算符
                row_data['Params'] = sheet.cell(i, 6).value.replace('tel',str(tel))#替换值tel
                self.updata_tel(int(tel)+1)
            else:#如果没有tel这个子字符串  就不需要去替换了
                row_data['Params'] = sheet.cell(i, 6).value
            row_data['ExpectedResult'] = sheet.cell(i, 7).value
            test_data.append(row_data)
        wb.close()
        final_data=[]  # 空列表 存储最终的测试用例数据
        if case_id == 'all':  # 如果case_id==all 那就获取所有的用例数据
            final_data=test_data  # 把测试用例赋值给final_data这个变量
        else:  # 否则  如果是列表 那就获取列表里面指定id的用例数据
            for i in case_id:  # 遍历case_id 里面的值
                final_data.append(test_data[i - 1])  # ? 对应关系??
        return final_data

    def get_tel(self):
        '''获取存在Excel里面的手机号码'''
        wb = load_workbook(self.file_name)
        sheet = wb['tel']
        wb.close()
        return sheet.cell(1,2).value#

    def updata_tel(self,new_tel):
        '''写回手机号码'''
        wb = load_workbook(self.file_name)
        sheet = wb['tel']
        sheet.cell(1,2,new_tel)
        wb.save(self.file_name)
        wb.close()

    def write_back(self,row, col, value):
        '''写回测试结果到Excel中'''
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        sheet.cell(row, col).value = value
        wb.save(self.file_name)
        wb.close()


if __name__ == '__main__':
    sheet_name = 'recharge'
    test_data = DoExcel(project_path.case_path,sheet_name).read_data()
    print(test_data)
    # print(DoExcel(project_path.case_path,'tel').get_tel())
