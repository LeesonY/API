#-*- coding: utf-8 -*-
from openpyxl import workbook
from openpyxl import load_workbook
class DoExcel:
    '''类的作用是完成Excel数据的读写，新建表单的操作'''
    def fun_1(self):
        '''读取所有的数据，以嵌套列表的形式存储，每一行都是一个子列表，每一个单元格都是子列表里面的元素'''
        pass

    def write_excel(self,file_name,sheet_name,row,col,value):
        '''在指定的单元格写入指定的数据，并保存到当前的Excel'''
        wb=load_workbook(file_name)
        sheet=wb[sheet_name]
        sheet.cell(row,col).value=value
        wb.save(file_name)


    def create_excel(self,file_name,sheet_name):
        """新建一个Excel"""
        wb=workbook.Workbook()
        wb.create_sheet(sheet_name)
        wb.save(file_name)

if __name__ == '__main__':
    DoExcel().create_excel('python21.xlsx','lemon')
    DoExcel().write_excel('python_14.xlsx','Sheet1',1,2,'信python得永生!')